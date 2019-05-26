# -*- coding: utf-8 -*-
"""
Created on Thu Apr 18 13:53:22 2019

@author: miha.kerc


This script updates excel spreadsheets that cointain historical prices with new
OHLC data that it copies from another spreadsheet that has Thomson Reuters for
Excel enabled.
"""



import pandas as pd
from openpyxl import load_workbook
import datetime

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """


    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        print ("File not found")# file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    try:
        writer.save()
    except:
        print("Close the target file")
        return


#if __name__ == "__main__":
def update(markets, file_path, df_path):
    df = pd.read_excel(df_path, sheet_name="Data", index_col=0)
    today=datetime.date.today()
    today.strftime("%d.%m.%Y")
    has_changed = False
    # Values that get returned if market was closed today
    wrong_val = [0, None, float("nan"), ""]
    
    for n in range(len(markets)):
        cancel = False
        try:
            file = pd.read_excel(file_path, sheet_name=markets[n], index_col=1).tail(1)
        except Exception as e:
            print(e)
            continue
        
        df_o = round(df.iloc[n, 0],2)
        df_h = round(df.iloc[n, 1],2)
        df_l = round(df.iloc[n, 2],2)
        df_c = round(df.iloc[n, 3],2)
        
        file_o = round(file.iloc[0, 1],2)
        file_h = round(file.iloc[0, 2],2)
        file_l = round(file.iloc[0, 3],2)
        file_c = round(file.iloc[0, 4],2)
        
        
        # Check if the markets were closed today (Reuters returns 0 or NaN values)
        if df_o in wrong_val:
            cancel = True
        if df_h in wrong_val:
            cancel = True
        if df_l in wrong_val:
            cancel = True
        if df_c in wrong_val:
            cancel = True
        if(df_o == file_o and df_h == file_h and df_l == file_l and df_c == file_c):
            cancel = True
			
        # Update the file if there has been an update to the data
        if not(cancel):
            res={"Date":[today], "Open":df_o, "High":df_h, "Low":df_l, "Close":df_c} # make a dictionary of current values
            res=pd.DataFrame(data=res) # convert to Pandas dataframe
            print(markets[n],res)
            append_df_to_excel(file_path, res, markets[n], header=None) # add to excel
            has_changed = True # signal to master.py that data has changed
        else:
            print(markets[n]+":","No update")

    print("\n")
    return has_changed